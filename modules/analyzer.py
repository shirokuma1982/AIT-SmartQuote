from openpyxl import load_workbook
from openpyxl import Workbook


def create_data_dictionary(input_file, output_file):

    wb = load_workbook(input_file, read_only=True)
    ws = wb.active

    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    sample = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

    new_wb = Workbook()
    new_ws = new_wb.active

    new_ws.title = "データ辞書"

    new_ws.append([
        "No",
        "項目名",
        "サンプル値",
        "Version1使用",
        "備考"
    ])

    for i, header in enumerate(headers, start=1):

        value = ""

        if i <= len(sample):
            value = sample[i-1]

        new_ws.append([
            i,
            header,
            str(value),
            "",
            ""
        ])

        new_wb.save(output_file)
    new_wb.close()
    wb.close()

    print("=" * 35)
    print("データ辞書を作成しました！")
    print(output_file)
    print("=" * 35)