from openpyxl import load_workbook


def analyze_excel(file_path):

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    print("=" * 80)
    print("OceanQuote AI - Excel Analyzer")
    print("=" * 80)

    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    print("\n【ヘッダー一覧】")

    for i, header in enumerate(headers, start=1):
        print(f"{i:3} : {header}")

    print("\n" + "=" * 80)
    print("【データサンプル（2～6行目）】")
    print("=" * 80)

    for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        print(row)